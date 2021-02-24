import pandas as pd
import openpyxl
from itertools import chain
import sys

def checkNulls(val):
    if val.value is None:
        return ' '
    else:
        return str(val.value)


def main():
    mrPath = sys.argv[2] #'E:/BethelPark-NeilArmstrong-Meter Request Form.Updated5.4.xlsx'
    templatePath = 'template.xlsx'
    ovfPath = sys.argv[4] #'E:/MI-WO-OVF-24 Bethel Park School District - Neil Armstrong 272434 02-08-2021.xlsx'

    templateWb = openpyxl.load_workbook(templatePath,data_only = True)
    template = templateWb.active

    requestForm = openpyxl.load_workbook(mrPath,data_only = True).active
    ovf = openpyxl.load_workbook(ovfPath,data_only = True)['IDR 1']

    concatenated = chain(range(1,4), range(6, 16))
    for i in concatenated:
        cellNumber = 'B' + str(i)
        if i<5:
            template[cellNumber] = requestForm.cell(row = i+3, column = 4).value
        else:
            template[cellNumber] = requestForm.cell(row = i+2, column = 4).value


    template['B18'] = checkNulls(requestForm['C28'])
    template['B19'] = checkNulls(requestForm['D28'])
    template['B22'] = template['B23'] = checkNulls(requestForm['E28'])


    for i in range(66,71):
        template[chr(i) + '25'] = checkNulls(ovf[chr(i) + '55']).replace(' ','')
        template[chr(i) + '26'] = checkNulls(ovf[chr(i) + '54'])
        template[chr(i) + '27'] = checkNulls(ovf[chr(i) + '6'])
        template[chr(i) + '29'] = checkNulls(ovf[chr(i) + '29'])
        template[chr(i) + '30'] = checkNulls(ovf[chr(i) + '154'])
        template[chr(i) + '31'] = checkNulls(ovf[chr(i) + '27'])
        template[chr(i) + '32'] = checkNulls(ovf[chr(i) + '122'])
        template[chr(i) + '34'] = checkNulls(ovf[chr(i) + '60']) + ' ' + checkNulls(ovf[chr(i) + '62'])
        if all(x is not None for x in (ovf[chr(i) + '56'].value,ovf[chr(i) + '57'].value)):
            template[chr(i) + '35'] = checkNulls(ovf[chr(i) + '56']) + '/' + checkNulls(ovf[chr(i) + '57'])
        


    template['B38'] = checkNulls(ovf['B157'])
    template['B40'] = checkNulls(ovf['B158'])
    template['B41'] = checkNulls(ovf['B4'])
    template['B42'] = checkNulls(ovf['B40'])
    template['B43'] = checkNulls(ovf['B154']) + ' ' + checkNulls(ovf['B155'])
    template['B46'] = checkNulls(ovf['B154'])

    dest_filename = 'C:/Users/Public/Documents/InstallCompletionForm_' + checkNulls(requestForm['D12']).replace(' ','') +'.xlsx'
    templateWb.save(filename = dest_filename)
    
    print(dest_filename)


if __name__ == '__main__':
    main()